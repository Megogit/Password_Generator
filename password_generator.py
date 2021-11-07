import random
import string
import time
from datetime import date, datetime
from os import path

from openpyxl import Workbook, load_workbook

chars = ''
rpassword=[]
sizet= 0
generated_password=''
try:


    if path.exists('db.xlsx'):
        pass

    else:
        wb = Workbook()
        ws =  wb.active
        ws.title = "KAYIT"
        ws.cell(row = 1 ,column = 1).value = "KAYIT TARİHİ"
        ws.cell(row = 1 ,column = 2).value = "ŞİFRENİN KULLANILDIĞI YER"
        ws.cell(row = 1 ,column = 3).value = "ÜRETİLEN ŞİFRE"

        wb.save('db.xlsx')
        row = ws.max_row+1

    password_place = input("şifreyi kullanacağın yer?")
    while True:
        try:
            size_u = int(input("şifre kaç haneli olacak ?"))
            break
        except:
            print("lütfen sayı değeri giriniz")
    while True:
        upper_case = input("Büyük harf olacak mı (y/n) ?").lower()
        if upper_case == "y" or upper_case== "n":
            break
        else:
            print("lütfen cevabınız y yada n olsun")

    while True:
        lower_case = input("küçük harf olacak mı (y/n) ?").lower()
        if lower_case == "y" or lower_case == "n":
            break
        else:
            print("lütfen cevabınız y yada n olsun")


    while True:
        number = input("sayı olacak mı olacak mı (y/n) ?").lower()
        if number == "y" or number == "n":
            break
        else:
            print("lütfen cevabınız y yada n olsun")

    while True:
        punct = input("noktalama işareti olacak mı (y/n) ?").lower()
        if punct == "y" or punct == "n":
            break
        else:
            print("lütfen cevabınız y yada n olsun")



    if upper_case == 'y':
        rpassword.append( random.choice(string.ascii_uppercase))
    if lower_case == 'y':
        rpassword.append(random.choice(string.ascii_lowercase))
    if number =='y':
        rpassword.append(random.choice(string.digits))
    if punct =='y':
        rpassword.append(random.choice(string.punctuation))

    password_lengt = len(rpassword)

    if password_lengt > size_u:
        raise Exception("Üzgünüm ama bu şartları sağlayan bir password oluşturulamaz. Programı yeniden başlat")

    up = string.ascii_uppercase if upper_case == 'y' else ''
    low = string.ascii_lowercase if lower_case == 'y' else ''
    num = string.digits if number == 'y' else ''
    punc = string.punctuation if punct == 'y' else ''

    charst= up + low+ num + punc

    sizet =size_u - password_lengt






    def pass_generator(size= sizet, chars=charst):
        return ''.join(random.choice(chars) for _ in range(size))

    try:
        generated_password =pass_generator()+"".join(str(x) for x in rpassword)
        print(generated_password)

    except:
        print (" arkadaşım şifre dediğinde harf marf birşey olur... adam akıllı bir şey iste")



    wb = load_workbook("db.xlsx")
    ws = wb['KAYIT']
    row = ws.max_row+1
    ws.cell(row=row, column=1).value = datetime.now()
    ws.cell(row=row, column=2).value = password_place
    ws.cell(row=row, column=3).value = generated_password
    wb.save('db.xlsx')

    wb.close()
except Exception as e: print(e)


input()
